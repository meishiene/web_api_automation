def test_web_test_case_crud_for_owner(client, create_user_and_login, auth_headers):
    token = create_user_and_login("owner_web", "pwd")
    headers = auth_headers(token)

    project_resp = client.post(
        "/api/projects/",
        headers=headers,
        json={"name": "P-web", "description": "desc"},
    )
    assert project_resp.status_code == 200
    project_id = project_resp.json()["id"]

    create_resp = client.post(
        "/api/web-test-cases",
        headers=headers,
        json={
            "project_id": project_id,
            "name": "login-smoke",
            "description": "basic login flow",
            "base_url": "https://example.com",
            "browser_name": "firefox",
            "viewport_width": 1366,
            "viewport_height": 768,
            "timeout_ms": 45000,
            "headless": False,
            "capture_on_failure": True,
            "record_video": True,
            "steps": [
                {"action": "open", "params": {"url": "/login"}},
                {"action": "click", "params": {"selector": "#submit"}},
            ],
        },
    )
    assert create_resp.status_code == 200
    created = create_resp.json()
    assert created["project_id"] == project_id
    assert created["name"] == "login-smoke"
    assert created["browser_name"] == "firefox"
    assert created["viewport_width"] == 1366
    assert created["viewport_height"] == 768
    assert created["timeout_ms"] == 45000
    assert created["headless"] is False
    assert created["capture_on_failure"] is True
    assert created["record_video"] is True
    assert len(created["steps"]) == 2
    assert created["steps"][0]["order_index"] == 0
    assert created["steps"][1]["order_index"] == 1

    case_id = created["id"]

    list_resp = client.get(f"/api/web-test-cases/project/{project_id}", headers=headers)
    assert list_resp.status_code == 200
    assert len(list_resp.json()) == 1

    detail_resp = client.get(f"/api/web-test-cases/{case_id}", headers=headers)
    assert detail_resp.status_code == 200
    detail = detail_resp.json()
    assert detail["id"] == case_id
    assert detail["name"] == "login-smoke"
    assert len(detail["steps"]) == 2

    update_resp = client.put(
        f"/api/web-test-cases/{case_id}",
        headers=headers,
        json={
            "name": "login-smoke-updated",
            "description": "updated",
            "base_url": "https://example.com",
            "browser_name": "webkit",
            "viewport_width": 375,
            "viewport_height": 667,
            "timeout_ms": 15000,
            "headless": True,
            "capture_on_failure": False,
            "record_video": False,
            "steps": [
                {"action": "open", "params": {"url": "/login"}},
                {"action": "input", "params": {"selector": "#username", "value": "demo"}},
                {"action": "click", "params": {"selector": "#submit"}},
            ],
        },
    )
    assert update_resp.status_code == 200
    updated = update_resp.json()
    assert updated["name"] == "login-smoke-updated"
    assert updated["browser_name"] == "webkit"
    assert updated["viewport_width"] == 375
    assert updated["viewport_height"] == 667
    assert updated["timeout_ms"] == 15000
    assert updated["headless"] is True
    assert updated["capture_on_failure"] is False
    assert updated["record_video"] is False
    assert len(updated["steps"]) == 3
    assert updated["steps"][2]["order_index"] == 2

    delete_resp = client.delete(f"/api/web-test-cases/{case_id}", headers=headers)
    assert delete_resp.status_code == 200
    assert delete_resp.json()["message"] == "Web test case deleted"


def test_web_test_case_duplicate_name_rejected(client, create_user_and_login, auth_headers):
    token = create_user_and_login("owner_web_dup", "pwd")
    headers = auth_headers(token)

    project_resp = client.post(
        "/api/projects/",
        headers=headers,
        json={"name": "P-web-dup", "description": "desc"},
    )
    project_id = project_resp.json()["id"]

    first = client.post(
        "/api/web-test-cases",
        headers=headers,
        json={"project_id": project_id, "name": "dup", "steps": [{"action": "open", "params": {"url": "/"}}]},
    )
    assert first.status_code == 200

    duplicate = client.post(
        "/api/web-test-cases",
        headers=headers,
        json={"project_id": project_id, "name": "dup", "steps": [{"action": "open", "params": {"url": "/"}}]},
    )
    assert duplicate.status_code == 400
    body = duplicate.json()
    assert body["error"]["code"] == "WEB_TEST_CASE_ALREADY_EXISTS"


def test_owner_can_copy_web_test_case_with_steps(client, create_user_and_login, auth_headers):
    token = create_user_and_login("owner_web_copy", "pwd")
    headers = auth_headers(token)

    project_id = client.post(
        "/api/projects/",
        headers=headers,
        json={"name": "P-web-copy", "description": "desc"},
    ).json()["id"]

    create_resp = client.post(
        "/api/web-test-cases",
        headers=headers,
        json={
            "project_id": project_id,
            "name": "copy-source",
            "description": "source case",
            "base_url": "https://example.com",
            "browser_name": "firefox",
            "viewport_width": 1366,
            "viewport_height": 768,
            "timeout_ms": 45000,
            "headless": False,
            "capture_on_failure": True,
            "record_video": True,
            "steps": [
                {"action": "open", "params": {"url": "/login"}},
                {"action": "input", "params": {"locator_type": "css", "locator": "#username", "value": "demo"}},
                {"action": "click", "params": {"locator_type": "css", "locator": "#submit"}},
            ],
        },
    )
    assert create_resp.status_code == 200
    source = create_resp.json()

    copy_resp = client.post(f"/api/web-test-cases/{source['id']}/copy", headers=headers, json={})
    assert copy_resp.status_code == 200
    copied = copy_resp.json()
    assert copied["id"] != source["id"]
    assert copied["name"] == "copy-source-copy"
    assert copied["description"] == source["description"]
    assert copied["base_url"] == source["base_url"]
    assert copied["browser_name"] == source["browser_name"]
    assert copied["viewport_width"] == source["viewport_width"]
    assert copied["viewport_height"] == source["viewport_height"]
    assert copied["timeout_ms"] == source["timeout_ms"]
    assert copied["headless"] == source["headless"]
    assert copied["capture_on_failure"] == source["capture_on_failure"]
    assert copied["record_video"] == source["record_video"]
    assert len(copied["steps"]) == 3
    assert copied["steps"][0]["params"]["url"] == "/login"
    assert copied["steps"][1]["params"]["value"] == "demo"
    assert copied["steps"][2]["params"]["locator"] == "#submit"


def test_non_owner_cannot_create_web_test_case_on_foreign_project(client, create_user_and_login, auth_headers):
    owner_token = create_user_and_login("owner_web_foreign", "pwd")
    attacker_token = create_user_and_login("attacker_web_foreign", "pwd")

    project_resp = client.post(
        "/api/projects/",
        headers=auth_headers(owner_token),
        json={"name": "P-foreign-web", "description": "desc"},
    )
    project_id = project_resp.json()["id"]

    create_resp = client.post(
        "/api/web-test-cases",
        headers=auth_headers(attacker_token),
        json={"project_id": project_id, "name": "steal", "steps": [{"action": "open", "params": {"url": "/"}}]},
    )
    assert create_resp.status_code == 403
    assert create_resp.json()["error"]["code"] == "FORBIDDEN"


def test_owner_can_bulk_delete_web_test_cases(client, create_user_and_login, auth_headers):
    token = create_user_and_login("owner_web_bulk_delete", "pwd")
    headers = auth_headers(token)

    project_id = client.post(
        "/api/projects/",
        headers=headers,
        json={"name": "P-web-bulk-delete", "description": "desc"},
    ).json()["id"]

    case1_id = client.post(
        "/api/web-test-cases",
        headers=headers,
        json={"project_id": project_id, "name": "web-bulk-case-1", "steps": [{"action": "open", "params": {"url": "/"}}]},
    ).json()["id"]
    case2_id = client.post(
        "/api/web-test-cases",
        headers=headers,
        json={"project_id": project_id, "name": "web-bulk-case-2", "steps": [{"action": "open", "params": {"url": "/login"}}]},
    ).json()["id"]

    bulk_delete_resp = client.post(
        f"/api/web-test-cases/project/{project_id}/bulk-delete",
        headers=headers,
        json={"test_case_ids": [case1_id, case2_id]},
    )
    assert bulk_delete_resp.status_code == 200
    body = bulk_delete_resp.json()
    assert body["deleted_count"] == 2
    assert body["deleted_ids"] == [case1_id, case2_id]

    list_resp = client.get(f"/api/web-test-cases/project/{project_id}", headers=headers)
    assert list_resp.status_code == 200
    assert list_resp.json() == []


def test_web_test_case_template_and_export_excel(client, create_user_and_login, auth_headers):
    token = create_user_and_login("owner_web_excel_export", "pwd")
    headers = auth_headers(token)

    project_id = client.post(
        "/api/projects/",
        headers=headers,
        json={"name": "P-web-excel-export", "description": "desc"},
    ).json()["id"]

    create_resp = client.post(
        "/api/web-test-cases",
        headers=headers,
        json={
            "project_id": project_id,
            "name": "excel-export-case",
            "steps": [
                {"action": "open", "params": {"url": "/login"}},
                {"action": "click", "params": {"locator_type": "css", "locator": "#submit"}},
            ],
        },
    )
    assert create_resp.status_code == 200

    template_resp = client.get(f"/api/web-test-cases/project/{project_id}/template.xlsx", headers=headers)
    assert template_resp.status_code == 200
    template_book = load_workbook(filename=BytesIO(template_resp.content))
    assert "Web Cases" in template_book.sheetnames
    assert "字段说明 Field Guide" in template_book.sheetnames
    assert template_book["Web Cases"]["A1"].value == "用例ID / Case ID"

    export_resp = client.get(f"/api/web-test-cases/project/{project_id}/export.xlsx", headers=headers)
    assert export_resp.status_code == 200
    export_book = load_workbook(filename=BytesIO(export_resp.content))
    data_sheet = export_book["Web Cases"]
    assert data_sheet["B2"].value == "excel-export-case"
    assert data_sheet["M2"].value == "open"
    assert data_sheet["N2"].value == "/login"
    assert data_sheet["M3"].value == "click"
    assert data_sheet["P3"].value == "#submit"


def test_web_test_case_import_excel_can_create_and_update(client, create_user_and_login, auth_headers):
    token = create_user_and_login("owner_web_excel_import", "pwd")
    headers = auth_headers(token)

    project_id = client.post(
        "/api/projects/",
        headers=headers,
        json={"name": "P-web-excel-import", "description": "desc"},
    ).json()["id"]

    existing_case = client.post(
        "/api/web-test-cases",
        headers=headers,
        json={
            "project_id": project_id,
            "name": "existing-case",
            "description": "before",
            "steps": [{"action": "open", "params": {"url": "/before"}}],
        },
    ).json()

    workbook_rows = build_template_rows()
    workbook_rows[0][0] = existing_case["id"]
    workbook_rows[0][1] = "existing-case"
    workbook_rows[0][2] = "updated by excel"
    workbook_rows[0][13] = "/updated-login"
    workbook_rows[1][0] = existing_case["id"]
    workbook_rows[1][1] = "existing-case"
    workbook_rows[2][0] = existing_case["id"]
    workbook_rows[2][1] = "existing-case"
    workbook_rows[3][0] = existing_case["id"]
    workbook_rows[3][1] = "existing-case"
    workbook_rows.extend(
        [
            ["", "new-imported-case", "created from excel", "https://example.com", "chromium", 1280, 720, 20000, True, True, False, 1, "open", "/new", "", "", "", "", "", ""],
            ["", "new-imported-case", "created from excel", "https://example.com", "chromium", 1280, 720, 20000, True, True, False, 2, "assert", "", "text", "创建成功", "", "创建成功", "", ""],
        ]
    )
    workbook_stream = build_workbook(workbook_rows)
    payload = {
        "file_name": "web-import.xlsx",
        "file_content_base64": b64encode(workbook_stream.getvalue()).decode("ascii"),
    }

    import_resp = client.post(
        f"/api/web-test-cases/project/{project_id}/import/xlsx",
        headers=headers,
        json=payload,
    )
    assert import_resp.status_code == 200
    body = import_resp.json()
    assert body["imported"] == 1
    assert body["updated"] == 1
    assert body["updated_case_ids"] == [existing_case["id"]]
    assert len(body["created_case_ids"]) == 1

    list_resp = client.get(f"/api/web-test-cases/project/{project_id}", headers=headers)
    assert list_resp.status_code == 200
    cases = {item["name"]: item for item in list_resp.json()}
    assert cases["existing-case"]["description"] == "updated by excel"
    assert len(cases["existing-case"]["steps"]) == 4
    assert cases["existing-case"]["steps"][0]["params"]["url"] == "/updated-login"
    assert cases["new-imported-case"]["description"] == "created from excel"
    assert len(cases["new-imported-case"]["steps"]) == 2
from base64 import b64encode
from io import BytesIO

from openpyxl import load_workbook

from app.services.web_test_case_excel import build_template_rows, build_workbook
