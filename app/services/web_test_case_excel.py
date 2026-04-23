import re
from base64 import b64decode
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

from app.errors import AppException, ErrorCode

WORKBOOK_SHEET_NAME = "Web Cases"
GUIDE_SHEET_NAME = "字段说明 Field Guide"
ALLOWED_ACTIONS = {"open", "click", "input", "wait", "assert", "screenshot"}
HEADER_PATTERN = re.compile(r"[\s_/()\-\n]+")

WEB_CASE_COLUMNS = [
    ("case_id", "用例ID", "Case ID", False, "可选。用于按 ID 更新现有用例。", "12"),
    ("case_name", "用例名称", "Case Name", True, "项目内唯一；未填 ID 时按名称匹配或新建。", "登录流程 smoke"),
    ("description", "描述", "Description", False, "用例说明。", "验证登录成功"),
    ("base_url", "基础地址", "Base URL", False, "站点基础地址。", "https://example.com"),
    ("browser_name", "浏览器", "Browser", False, "chromium / firefox / webkit。", "chromium"),
    ("viewport_width", "视口宽度", "Viewport Width", False, "浏览器窗口宽度。", "1366"),
    ("viewport_height", "视口高度", "Viewport Height", False, "浏览器窗口高度。", "768"),
    ("timeout_ms", "超时毫秒", "Timeout(ms)", False, "单用例超时，默认 30000。", "30000"),
    ("headless", "无头模式", "Headless", False, "true/false、1/0、是/否。", "true"),
    ("capture_on_failure", "失败截图", "Capture On Failure", False, "失败时是否截图。", "true"),
    ("record_video", "录制视频", "Record Video", False, "执行时是否录制视频。", "false"),
    ("step_order", "步骤序号", "Step Order", False, "步骤顺序；为空按行顺序导入。", "1"),
    ("action", "动作", "Action", False, "open / click / input / wait / assert / screenshot。", "open"),
    ("open_url", "打开地址", "Open URL", False, "仅 open 使用。", "/login"),
    ("locator_strategy", "定位方式", "Locator Strategy", False, "css / xpath / text / testid / role。", "css"),
    ("locator", "定位内容", "Locator", False, "元素定位表达式。", "#submit"),
    ("input_value", "输入值", "Input Value", False, "仅 input 使用。", "demo-user"),
    ("expected_text", "期望文本", "Expected Text", False, "仅 assert 使用。", "欢迎回来"),
    ("wait_ms", "等待毫秒", "Wait(ms)", False, "仅 wait 使用；默认 500。", "1000"),
    ("screenshot_filename", "截图文件名", "Screenshot Filename", False, "仅 screenshot 使用。", "login-success.png"),
]


def header_label(column: tuple[str, str, str, bool, str, str]) -> str:
    return f"{column[1]} / {column[2]}"


def normalize_header_text(value: Any) -> str:
    return HEADER_PATTERN.sub("", str(value or "").strip().lower())


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    value = str(value).strip()
    return value or None


def parse_int(value: Any, field_name: str, row_number: int) -> int | None:
    normalized = normalize_text(value)
    if normalized is None:
        return None
    try:
        return int(float(normalized))
    except ValueError as exc:
        raise AppException(400, ErrorCode.VALIDATION_ERROR, f"Invalid integer for {field_name} at row {row_number}") from exc


def parse_bool(value: Any, field_name: str, row_number: int) -> bool | None:
    normalized = normalize_text(value)
    if normalized is None:
        return None
    lowered = normalized.lower()
    if lowered in {"true", "1", "yes", "y", "是"}:
        return True
    if lowered in {"false", "0", "no", "n", "否"}:
        return False
    raise AppException(400, ErrorCode.VALIDATION_ERROR, f"Invalid boolean for {field_name} at row {row_number}")


def build_workbook(case_rows: list[list[Any]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = WORKBOOK_SHEET_NAME
    sheet.freeze_panes = "A2"
    sheet.append([header_label(column) for column in WEB_CASE_COLUMNS])
    for row in case_rows:
        sheet.append(row)

    guide_sheet = workbook.create_sheet(GUIDE_SHEET_NAME)
    guide_sheet.freeze_panes = "A2"
    guide_sheet.append(["字段 / Field", "中文名 / Chinese", "英文名 / English", "是否必填 / Required", "说明 / Description", "示例 / Example"])
    for key, header_cn, header_en, required, description, example in WEB_CASE_COLUMNS:
        guide_sheet.append([key, header_cn, header_en, "是 / Yes" if required else "否 / No", description, example])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_template_rows() -> list[list[Any]]:
    return [
        ["", "登录流程 smoke", "登录页输入账号密码后进入首页", "https://example.com", "chromium", 1366, 768, 30000, True, True, False, 1, "open", "/login", "", "", "", "", "", ""],
        ["", "登录流程 smoke", "登录页输入账号密码后进入首页", "https://example.com", "chromium", 1366, 768, 30000, True, True, False, 2, "input", "", "css", "#username", "demo-user", "", "", ""],
        ["", "登录流程 smoke", "登录页输入账号密码后进入首页", "https://example.com", "chromium", 1366, 768, 30000, True, True, False, 3, "click", "", "css", "#submit", "", "", "", ""],
        ["", "登录流程 smoke", "登录页输入账号密码后进入首页", "https://example.com", "chromium", 1366, 768, 30000, True, True, False, 4, "assert", "", "text", "欢迎回来", "", "欢迎回来", "", ""],
    ]


def build_rows_from_cases(cases: list[Any], parse_params_fn) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for case in cases:
        steps = case.steps or [None]
        for step in steps:
            params = parse_params_fn(step.params_json) if step else {}
            rows.append(
                [
                    case.id,
                    case.name,
                    case.description,
                    case.base_url,
                    case.browser_name,
                    case.viewport_width,
                    case.viewport_height,
                    case.timeout_ms,
                    bool(case.headless),
                    bool(case.capture_on_failure),
                    bool(case.record_video),
                    (step.order_index + 1) if step else None,
                    step.action if step else None,
                    params.get("url"),
                    params.get("locator_type"),
                    params.get("locator"),
                    params.get("value"),
                    params.get("contains"),
                    params.get("timeout_ms"),
                    params.get("filename"),
                ]
            )
    return rows


def load_rows_from_base64(file_content_base64: str) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(filename=BytesIO(b64decode(file_content_base64)))
    except Exception as exc:
        raise AppException(400, ErrorCode.VALIDATION_ERROR, "Invalid Excel workbook content") from exc

    sheet = workbook[WORKBOOK_SHEET_NAME] if WORKBOOK_SHEET_NAME in workbook.sheetnames else workbook.worksheets[0]
    headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if headers is None:
        raise AppException(400, ErrorCode.VALIDATION_ERROR, "Workbook must contain a header row")

    raw_header_map = {normalize_header_text(value): index for index, value in enumerate(headers) if value is not None}
    column_map: dict[str, int] = {}
    for key, header_cn, header_en, _required, _description, _example in WEB_CASE_COLUMNS:
        variants = {key, header_cn, header_en, f"{header_cn} / {header_en}"}
        for variant in variants:
            normalized = normalize_header_text(variant)
            if normalized in raw_header_map:
                column_map[key] = raw_header_map[normalized]
                break
        else:
            raise AppException(400, ErrorCode.VALIDATION_ERROR, f"Workbook missing column: {header_cn} / {header_en}")

    rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        item = {key: row[index] if index < len(row) else None for key, index in column_map.items()}
        if all(normalize_text(value) is None for value in item.values()):
            continue
        item["_row_number"] = row_number
        rows.append(item)

    if not rows:
        raise AppException(400, ErrorCode.VALIDATION_ERROR, "Workbook contains no importable rows")
    return rows
